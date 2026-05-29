"""Stage 5 — write a list of Transactions to a .xlsx file in the output directory.

Column order mirrors the canonical accounting schema: Date, Description,
Money Out, Money In, Balance. Rows with confidence="low" receive a "?" marker
in a Notes column so a human reviewer can spot OCR errors quickly.
"""

from __future__ import annotations

import logging
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]

from statement_to_excel.models import Reconciliation, Statement, Transaction

log = logging.getLogger(__name__)

_HEADERS = ("Date", "Description", "Money Out", "Money In", "Balance", "Notes")


def export(statement: Statement, output_dir: Path) -> Path:
    """Serialise a Statement to a .xlsx file.

    Args:
        statement: Validated transactions plus source metadata.
        output_dir: Destination directory; created if absent.

    Returns:
        Path to the written .xlsx file.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{statement.source_pdf.stem}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(list(_HEADERS))

    for txn in statement.transactions:
        ws.append(_row(txn))

    # Light formatting so the file is pleasant to open without manual tweaking.
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 48
    for col in ("C", "D", "E"):
        ws.column_dimensions[col].width = 14
        for cell in ws[col][1:]:
            cell.number_format = "#,##0.00"
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"

    if statement.reconciliation is not None:
        _write_reconciliation_sheet(wb, statement.reconciliation)

    wb.save(out_path)
    log.info("export: wrote %d row(s) to %s", len(statement.transactions), out_path)
    return out_path


def _write_reconciliation_sheet(wb: Workbook, rec: Reconciliation) -> None:
    """Add a "Summary" sheet showing stated-vs-extracted totals and a verdict.

    Gives the human reviewer the accountant's check at a glance: a PASS, or a
    REVIEW with the specific discrepancies spelled out.
    """
    ws = wb.create_sheet("Summary")
    computed_closing = (
        rec.opening_balance + rec.extracted_paid_in - rec.extracted_paid_out
        if rec.opening_balance is not None
        else None
    )
    ws.append(["Reconciliation", "PASS" if rec.ok else "REVIEW"])
    ws.append([])
    ws.append(["", "Statement", "Extracted"])
    ws.append(["Payments In", _to_excel(rec.stated_paid_in), _to_excel(rec.extracted_paid_in)])
    ws.append(["Payments Out", _to_excel(rec.stated_paid_out), _to_excel(rec.extracted_paid_out)])
    ws.append(["Opening balance", _to_excel(rec.opening_balance), ""])
    ws.append(["Closing balance", _to_excel(rec.closing_balance), _to_excel(computed_closing)])
    if rec.issues:
        ws.append([])
        ws.append(["Issues"])
        for issue in rec.issues:
            ws.append([issue])
    ws.column_dimensions["A"].width = 18
    for col in ("B", "C"):
        ws.column_dimensions[col].width = 16


def _row(t: Transaction) -> list[object]:
    return [
        t.date,
        t.description,
        _to_excel(t.money_out),
        _to_excel(t.money_in),
        _to_excel(t.balance),
        "?" if t.confidence == "low" else "",
    ]


def _to_excel(value: Decimal | None) -> float | None:
    """Convert Decimal to float at the spreadsheet boundary.

    Excel stores numbers as IEEE-754 doubles, so the Decimal precision we
    maintain inside the pipeline ends here. Two-decimal money values are well
    inside float resolution.
    """
    return float(value) if value is not None else None
