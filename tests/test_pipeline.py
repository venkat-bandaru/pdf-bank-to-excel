"""End-to-end test: feed the Jan26 sample through pipeline.run() and verify
the resulting .xlsx has the expected number of rows and a valid balance chain.
"""

from __future__ import annotations

import shutil
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from statement_to_excel import pipeline
from statement_to_excel.models import Config

# Hand-counted from tests/samples/Jan26.pdf (5 pages of TSB statement).
_EXPECTED_ROWS = 41
_TOLERANCE = Decimal("0.01")


@pytest.fixture
def jan26_config(tmp_path: Path, samples_dir: Path) -> Config:
    """Stage Jan26.pdf in a per-test input directory and return a matching Config."""
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    failed_dir = tmp_path / "failed"
    log_dir = tmp_path / "logs"
    input_dir.mkdir()
    shutil.copy(samples_dir / "Jan26.pdf", input_dir / "Jan26.pdf")
    return Config(
        input_dir=input_dir,
        output_dir=output_dir,
        failed_dir=failed_dir,
        log_dir=log_dir,
        ocr_min_chars_per_page=100,
        ocr_tesseract_lang="eng",
        extractor_priority=("hsbc", "barclays", "generic"),
    )


def test_pipeline_jan26_end_to_end(jan26_config: Config) -> None:
    pipeline.run(jan26_config)

    out_path = jan26_config.output_dir / "Jan26.xlsx"
    assert out_path.is_file(), "expected pipeline to produce Jan26.xlsx"

    wb = load_workbook(out_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]

    assert header == ("Date", "Description", "Money Out", "Money In", "Balance", "Notes")
    assert len(data) == _EXPECTED_ROWS

    # Balance chain validation. Rows are in the same order as the PDF
    # (newest first), so for each pair (i, i+1) the chronologically prior
    # balance is at i+1 and the signed amount must close the gap.
    for i in range(len(data) - 1):
        cur_date, _, cur_out, cur_in, cur_bal, _ = data[i]
        _, _, _, _, prv_bal, _ = data[i + 1]
        delta = Decimal(str(cur_bal)) - Decimal(str(prv_bal))
        signed = Decimal(str(cur_in or 0)) - Decimal(str(cur_out or 0))
        assert abs(delta - signed) <= _TOLERANCE, (
            f"balance chain broken at row {i} ({cur_date}): "
            f"delta={delta}, signed={signed}"
        )
